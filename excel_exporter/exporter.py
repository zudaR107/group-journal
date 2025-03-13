# excel_exporter.py

import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from models.journal_manager import JournalManager

class ExcelExporter:
    """
    Класс для экспорта журнала посещаемости в Excel.
    """

    def __init__(self, journal_manager: JournalManager):
        self.journal_manager = journal_manager

    def export_to_excel(self, file_path: str) -> None:
        """
        Экспортирует данные в Excel-файл.
        :param file_path: Путь к файлу для сохранения.
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Удаляем стандартный пустой лист

        # Создаём лист для каждого предмета
        for subject in self.journal_manager.subjects:
            self._create_subject_sheet(wb, subject)

        # Добавляем вкладку "Пропущенные часы"
        self._create_summary_sheet(wb)

        # Сохраняем файл
        wb.save(file_path)

    def _create_subject_sheet(self, wb, subject) -> None:
        """
        Создаёт лист для предмета с таблицами по каждому типу занятий.
        """
        sheet = wb.create_sheet(title=subject.subject_name[:31])  # Название листа (до 31 символа)

        # Создаём заголовок
        sheet.append(["П. группа", "№", "Фамилия И.О. студента", "Учёт посещаемости лекций по дисциплине " + subject.subject_name])
        sheet.append(["", "", "", "№ недели"] + [str(i + 1) for i in range(18)])

        # Устанавливаем стили для заголовков
        for col in range(1, 23):  # A - W (максимально 22 колонки)
            sheet.cell(row=1, column=col).font = Font(bold=True)
            sheet.cell(row=2, column=col).font = Font(bold=True)
            sheet.cell(row=1, column=col).alignment = Alignment(horizontal="center")
            sheet.cell(row=2, column=col).alignment = Alignment(horizontal="center")

        # Определяем номер строки для вставки данных
        row_index = 3

        # Заполняем список студентов
        for idx, student in enumerate(self.journal_manager.group_info.students, start=1):
            subgroup_roman = self._to_roman(student.subgroup)  # Переводим подгруппу в римские цифры
            sheet.append([subgroup_roman, idx, student.full_name] + [""] * 18)
            row_index += 1

        # Настраиваем ширину колонок
        sheet.column_dimensions["A"].width = 5
        sheet.column_dimensions["B"].width = 5
        sheet.column_dimensions["C"].width = 25
        for col in range(4, 23):  # Настроить ширину всех недель
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 5

    def _create_summary_sheet(self, wb) -> None:
        """
        Создаёт вкладку "Пропущенные часы" с подсчётом пропусков по студентам.
        """
        sheet = wb.create_sheet(title="Пропущенные часы")
        sheet.append(["П. группа", "№", "Фамилия И. О.", "№ недели"] + [str(i + 1) for i in range(18)] + ["Всего часов"])

        # Заполняем таблицу студентами
        for idx, student in enumerate(self.journal_manager.group_info.students, start=1):
            subgroup_roman = self._to_roman(student.subgroup)  # Переводим подгруппу в римские цифры
            sheet.append([subgroup_roman, idx, student.full_name] + [""] * 18 + ["=SUM(E2:V2)"])

        # Настраиваем стили
        for col in sheet.iter_cols():
            for cell in col:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=12)

    def _to_roman(self, num: int) -> str:
        """
        Переводит номер подгруппы в римские цифры.
        """
        roman_numerals = {1: "I", 2: "II", 3: "III", 4: "IV", 5: "V", 6: "VI", 7: "VII", 8: "VIII", 9: "IX", 10: "X"}
        return roman_numerals.get(num, str(num))  # Если нет в списке — вернуть обычное число
